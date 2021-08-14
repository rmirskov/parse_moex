from selenium.webdriver import Firefox
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from functools import reduce
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import pandas as pd
import numpy as np
import random
import time

from send_message import send_mail, say_string
import settings

options = Options()
options.add_argument("-headless")
proxy = settings.PROXY[random.randint(0, len(settings.PROXY) - 1)]
options.add_argument(f'--proxy-server=%s{proxy}')

driver = Firefox('/home/darya/Документы/parse_moex', firefox_options=options)
driver.get('https://moex.com')

def find_and_click_elem(link):
        if '/' in link:
            elem = driver.find_element_by_xpath(link)
        else:
            elem = driver.find_element_by_class_name(link)
        elem.click()
        # drop modal window
        time.sleep(1)
        html = driver.find_element_by_tag_name('html') 
        html.send_keys(Keys.ESCAPE)

links = [
        "js-menu-dropdown-button",
        "//div[@class='item']//child::a[contains(text(), 'Срочный рынок')]",
        "/html/body/div[2]/div/div/div/div/div[1]/div/a[1]",
        "//div[contains(text(), 'Индикативные курсы')]//parent::span//parent::div//parent::a"
    ]

try:
    for link in links:
        find_and_click_elem(link)
    
    currency_pairs = ['USD_RUB', 'EUR_RUB']
    df_dict = {}
    for pair in currency_pairs:
        select = Select(driver.find_element_by_id('ctl00_PageContent_CurrencySelect'))
        select.select_by_value(pair)
        element = driver.find_element_by_class_name('tablels')
        num_row = len(element.find_elements_by_xpath('//tr[@*]'))
        values = [value.text.replace(',', '.') for value in element.find_elements_by_xpath('//tr[@*]//child::td')]
        num_col = int(len(values)/num_row)
        data_list = np.reshape(values, (num_row, num_col))
        df = pd.DataFrame(data_list)
        df.columns = [f'Дата {pair}', f'Курс ПК {pair}', f'Время ПК {pair}', f'Курс ОК {pair}', f'Время ОК {pair}']
        df = df.astype({f'Курс ПК {pair}': 'float64', f'Курс ОК {pair}': 'float64'})
        df[f'Дата {pair}'] = pd.to_datetime(df[f'Дата {pair}']).dt.date
        df[f'Изменение {pair}'] = df[f'Курс ОК {pair}'] - df[f'Курс ПК {pair}']
        df_dict[pair] = df[[f'Дата {pair}', f'Курс ОК {pair}', f'Изменение {pair}']]

    result_data = reduce(lambda x, y: x.join(y), [elem for elem in df_dict.values()])
    index_list = [i for i in range(4, len(result_data.iloc[0]), 3)]  # if currency_pairs relative RUB more then 2
    for i in index_list:
        name = result_data.columns[i-1].split(' ')[1].split('_')[0]
        result_data[f'Курс {name}_USD'] = result_data.iloc[:, i] / result_data.iloc[:, 1]
    sheet_name = time.ctime().replace(':', '-')
    file_name = 'data_sheet.xlsx'
    with pd.ExcelWriter(file_name, engine='openpyxl', date_format='dd.mm.yyyy') as writer:
        result_data.to_excel(writer, sheet_name=sheet_name, index=False, float_format="%.4f")

    wb = openpyxl.load_workbook(file_name)
    sheet = wb[sheet_name]
    date_col = [i + 1 for i in range(num_col) if i % 3 == 0]
    for i in range(1, len(result_data)+2):
        for j in range(1, len(result_data.iloc[0])+1):
            if i == 1:
                width = len(sheet.cell(i, j).value) * 1.3
                # sheet.column_dimensions[get_column_letter(j)].bestFit = True
                # sheet.column_dimensions[get_column_letter(j)].auto_size = True
                sheet.column_dimensions[get_column_letter(j)].width = width
            sheet.cell(i, j).alignment = Alignment(horizontal="center", vertical='center')
            if j not in date_col:
                if sheet.cell(1, j).value.split('_')[1] == 'RUB':
                    sheet.cell(i, j).number_format = '# ##0.0000" р.";-# ##0.0000" р."'
                elif sheet.cell(1, j).value.split('_')[1] == 'USD':
                    sheet.cell(i, j).number_format = '# ##0.0000" $";-# ##0.0000" $"'
    wb.save(file_name)

    mail_text = f'Создан отчет. В содержимом {num_row} {say_string(num_row)}.'
    send_mail('Робот', mail_text, 'rmir050607@gmail.com', file_name)

except Exception as err:
    print(err)

finally:
    driver.close()
    exit()
